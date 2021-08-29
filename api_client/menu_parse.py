import json
from openpyxl import load_workbook
from Parser import Parser


def menu_parse_schedule():
    file_name = input('Введите имя xlsx файла: ')
    organization = input('Введите название организации: ')
    header_row = int(input('Номер строки заголовков (по умолчанию 0): ') or '0')

    # Предварительная обработка файла (разбиение объединённых ячеек)
    wb = load_workbook(filename=file_name)

    for sheet in wb.worksheets:
        while sheet.merged_cells.ranges:
            for merge_range in sheet.merged_cells.ranges:
                (min_col, min_row, max_col, max_row) = merge_range.bounds
                sheet.unmerge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)

                fill_val = sheet.cell(column=min_col, row=min_row).value
                for col in range(min_col, max_col + 1):
                    for row in range(min_row, max_row + 1):
                        sheet.cell(column=col, row=row, value=fill_val)
    wb.save(file_name)

    # Парсинг файла с расписанием
    with Parser(file_name) as parser:
        data = parser.parse_schedule_from_excel(header_row)

    json_data = {organization: data}

    with open('schedule_data.json', 'w') as outfile:
        json.dump(json_data, outfile, indent=4, sort_keys=True)
    print('Расписание занятий сохранено в schedule_data.json')


def menu_parse_exams():
    file_name = input('Введите имя xls/xlsx файла: ')
    organization = input('Введите название организации: ')
    header_row = int(input('Номер строки заголовков (по умолчанию 0): '))

    with Parser(file_name) as parser:
        data = parser.parse_exams_from_excel(header_row)

    json_data = {organization: data}

    with open('exams_data.json', 'w') as outfile:
        json.dump(json_data, outfile, indent=4, sort_keys=True)
    print('Расписание экзаменов сохранено в exams_data.json')
